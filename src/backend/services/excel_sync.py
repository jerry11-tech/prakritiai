"""
Automatic Excel Synchronization Engine for PrakritiAI.
Maintains `Prakriti_Verified_Data.xlsx` in real time with 5 sheets:
  1. User_Data
  2. Verified_Data
  3. Change_History
  4. Verification_Log
  5. Summary
"""

import os
import sys
import sqlite3
from datetime import datetime
from typing import Any, Dict, List, Optional

# Ensure backend root is in sys.path
BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db.database import get_connection

EXCEL_FILE_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "Prakriti_Verified_Data.xlsx")
)


def format_header_row(ws, col_count: int, title: str = None) -> None:
    fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="medium", color="000000"))

    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28


def autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def sync_excel_file() -> Dict[str, Any]:
    """Reads SQLite database and overwrites/refreshes Prakriti_Verified_Data.xlsx."""
    sync_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # ----------------------------------------------------------------------
        # 1. Sheet 1: User_Data
        # ----------------------------------------------------------------------
        ws_user = wb.create_sheet(title="User_Data")
        user_headers = [
            "Participant ID",
            "Name",
            "Age Group",
            "Gender",
            "City",
            "Diabetes",
            "Blood Pressure",
            "Submission Date",
            "User Verification",
            "Verification Date",
            "Response Count",
            "Responses Breakdown",
        ]
        ws_user.append(user_headers)

        with get_connection() as conn:
            cursor = conn.cursor()
            participants = cursor.execute("""
                SELECT * FROM participants ORDER BY created_at DESC
            """).fetchall()

            verified_records_count = 0
            pending_records_count = 0
            needs_reverify_count = 0

            for p in participants:
                p_id = p["participant_id"]
                v_status = p["verification_status"] or "PENDING"
                if v_status == "VERIFIED":
                    verified_records_count += 1
                elif v_status == "NEEDS_REVERIFICATION":
                    needs_reverify_count += 1
                else:
                    pending_records_count += 1

                responses = cursor.execute("""
                    SELECT feature_key, feature_value FROM questionnaire_responses
                    WHERE participant_id = ?
                """, (p_id,)).fetchall()

                resp_str = "; ".join([f"{r['feature_key']}={r['feature_value']}" for r in responses])

                ws_user.append([
                    p_id,
                    p["name"] or "Anonymous",
                    p["age_group"] or "N/A",
                    p["gender"] or "N/A",
                    p["city"] or "N/A",
                    p["diabetes"] or "No",
                    p["blood_pressure"] or "Normal",
                    p["created_at"] or "",
                    v_status,
                    p["verification_date"] or "",
                    len(responses),
                    resp_str,
                ])

        format_header_row(ws_user, len(user_headers))
        autofit_columns(ws_user)

        # ----------------------------------------------------------------------
        # 2. Sheet 2: Verified_Data (ONLY user_verified = VERIFIED)
        # ----------------------------------------------------------------------
        ws_ver = wb.create_sheet(title="Verified_Data")
        ws_ver.append(user_headers)

        with get_connection() as conn:
            cursor = conn.cursor()
            verified_participants = cursor.execute("""
                SELECT * FROM participants WHERE verification_status = 'VERIFIED'
                ORDER BY verification_date DESC
            """).fetchall()

            for p in verified_participants:
                p_id = p["participant_id"]
                responses = cursor.execute("""
                    SELECT feature_key, feature_value FROM questionnaire_responses
                    WHERE participant_id = ?
                """, (p_id,)).fetchall()
                resp_str = "; ".join([f"{r['feature_key']}={r['feature_value']}" for r in responses])

                ws_ver.append([
                    p_id,
                    p["name"] or "Anonymous",
                    p["age_group"] or "N/A",
                    p["gender"] or "N/A",
                    p["city"] or "N/A",
                    p["diabetes"] or "No",
                    p["blood_pressure"] or "Normal",
                    p["created_at"] or "",
                    "VERIFIED",
                    p["verification_date"] or "",
                    len(responses),
                    resp_str,
                ])

        format_header_row(ws_ver, len(user_headers))
        autofit_columns(ws_ver)

        # ----------------------------------------------------------------------
        # 3. Sheet 3: Change_History (Complete Audit Trail)
        # ----------------------------------------------------------------------
        ws_chg = wb.create_sheet(title="Change_History")
        chg_headers = [
            "Change ID",
            "Participant ID",
            "Date & Time",
            "Changed By",
            "User Role",
            "Field / Question",
            "Previous Value",
            "New Value",
            "Change Type",
            "Reason",
            "Verification Status",
        ]
        ws_chg.append(chg_headers)

        with get_connection() as conn:
            cursor = conn.cursor()
            changes = cursor.execute("""
                SELECT * FROM change_history ORDER BY changed_at DESC
            """).fetchall()
            total_changes_count = len(changes)

            for c in changes:
                ws_chg.append([
                    c["change_id"],
                    c["participant_id"],
                    c["changed_at"],
                    c["changed_by"],
                    c["user_role"],
                    c["field_name"],
                    c["previous_value"],
                    c["new_value"],
                    c["change_type"],
                    c["reason"],
                    c["verification_status"],
                ])

        format_header_row(ws_chg, len(chg_headers))
        autofit_columns(ws_chg)

        # ----------------------------------------------------------------------
        # 4. Sheet 4: Verification_Log
        # ----------------------------------------------------------------------
        ws_vlog = wb.create_sheet(title="Verification_Log")
        vlog_headers = [
            "Verification ID",
            "Participant ID",
            "Verification Date",
            "Status",
            "Verified By",
            "Number of Answers",
            "Answers Changed Before Verification",
            "Verification Method",
        ]
        ws_vlog.append(vlog_headers)

        with get_connection() as conn:
            cursor = conn.cursor()
            vlogs = cursor.execute("""
                SELECT * FROM verification_log ORDER BY verification_date DESC
            """).fetchall()

            for v in vlogs:
                ws_vlog.append([
                    v["verification_id"],
                    v["participant_id"],
                    v["verification_date"],
                    v["status"],
                    v["verified_by"],
                    v["number_of_answers"],
                    v["answers_changed_before_verification"],
                    v["verification_method"],
                ])

        format_header_row(ws_vlog, len(vlog_headers))
        autofit_columns(ws_vlog)

        # ----------------------------------------------------------------------
        # 5. Sheet 5: Summary
        # ----------------------------------------------------------------------
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.column_dimensions["A"].width = 30
        ws_sum.column_dimensions["B"].width = 25

        ws_sum.append(["METRIC", "VALUE"])
        format_header_row(ws_sum, 2)

        today_str = datetime.now().strftime("%Y-%m-%d")
        with get_connection() as conn:
            cursor = conn.cursor()
            today_sub = cursor.execute("""
                SELECT COUNT(*) FROM participants WHERE created_at LIKE ?
            """, (f"{today_str}%",)).fetchone()[0]

            today_ver = cursor.execute("""
                SELECT COUNT(*) FROM participants WHERE verification_date LIKE ? AND verification_status = 'VERIFIED'
            """, (f"{today_str}%",)).fetchone()[0]

        summary_rows = [
            ("Total Submissions", len(participants)),
            ("Verified Users", verified_records_count),
            ("Pending Verification", pending_records_count),
            ("Needs Re-verification", needs_reverify_count),
            ("Total Changes Logged", total_changes_count),
            ("Today's Submissions", today_sub),
            ("Today's Verifications", today_ver),
            ("Last Excel Update", sync_time),
            ("Sync Status", "SUCCESS"),
        ]

        for label, val in summary_rows:
            ws_sum.append([label, val])

        # Save workbook
        wb.save(EXCEL_FILE_PATH)

        # Log sync event in DB
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO excel_sync_log (sync_id, sync_time, status, message, records_synced)
                VALUES (?, ?, ?, ?, ?)
            """, (f"SYNC_{int(datetime.now().timestamp())}", sync_time, "SUCCESS", "Synchronized successfully", len(participants)))
            conn.commit()

        return {
            "status": "SUCCESS",
            "syncTime": sync_time,
            "totalRecords": len(participants),
            "verifiedRecords": verified_records_count,
            "pendingRecords": pending_records_count,
            "needsReverificationRecords": needs_reverify_count,
            "totalChanges": total_changes_count,
            "todaySubmissions": today_sub,
            "todayVerifications": today_ver,
            "excelPath": EXCEL_FILE_PATH,
        }

    except Exception as e:
        error_msg = str(e)
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO excel_sync_log (sync_id, sync_time, status, message, records_synced)
                VALUES (?, ?, ?, ?, ?)
            """, (f"SYNC_ERR_{int(datetime.now().timestamp())}", sync_time, "FAILED", error_msg, 0))
            conn.commit()

        return {
            "status": "FAILED",
            "syncTime": sync_time,
            "error": error_msg,
        }
